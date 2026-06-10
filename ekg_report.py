from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ekg_translations import translate_columns


def _write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, **kwargs: Any) -> None:
    translate_columns(df).to_excel(writer, sheet_name=sheet_name, index=False, **kwargs)


def make_excel_report(
    scoring: dict[str, pd.DataFrame],
    inputs: pd.DataFrame,
    adverse: pd.DataFrame,
    fragments: pd.DataFrame,
    documents: pd.DataFrame,
    extraction_audit: pd.DataFrame,
    settings: dict[str, Any],
) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_df(writer, scoring["summary"], "Итог", startrow=3)
        _write_df(writer, inputs, "Полученные данные")
        _write_df(writer, adverse, "Негативные факты")
        _write_df(writer, scoring.get("missing_info", pd.DataFrame()), "Не найдено")
        _write_df(writer, scoring["ecology"], "Экология")
        _write_df(writer, scoring["cadres"], "Кадры")
        _write_df(writer, scoring["state_financial"], "Государство финансы")
        _write_df(writer, scoring["state_trust"], "Государство доверие")
        _write_df(writer, documents, "Документы")
        _write_df(writer, extraction_audit, "Журнал извлечения PDF")
        _write_df(writer, fragments, "Фрагменты")
        _write_df(writer, pd.DataFrame([settings]), "Настройки")

    buffer.seek(0)
    workbook = load_workbook(buffer)
    navy = PatternFill("solid", fgColor="0F4C5C")
    warning = PatternFill("solid", fgColor="FFF2CC")
    white_bold = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=15)

    dashboard = workbook["Итог"]
    dashboard.merge_cells("A1:C1")
    dashboard["A1"] = "Перевод GRI-отчета в ЭКГ-рейтинг"
    dashboard["A1"].fill = navy
    dashboard["A1"].font = title_font
    dashboard["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
    dashboard.merge_cells("A2:C2")
    dashboard["A2"] = "Расчёт выполнен по найденным и принятым значениям."
    dashboard["A2"].fill = warning
    dashboard["A2"].font = Font(bold=True, color="7F6000")
    dashboard["A2"].alignment = Alignment(wrap_text=True)

    for sheet in workbook.worksheets:
        header_row = 4 if sheet.title == "Итог" else 1
        sheet.freeze_panes = f"A{header_row+1}"
        for cell in sheet[header_row]:
            if cell.value is not None:
                cell.fill = navy
                cell.font = white_bold
                cell.alignment = Alignment(wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            max_width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            sheet.column_dimensions[letter].width = min(max(max_width + 2, 12), 60)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
